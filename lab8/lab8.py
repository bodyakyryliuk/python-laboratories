import argparse
import csv
import logging
import os
import re
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def run():
    parser = argparse.ArgumentParser(description="My application is designed to analyze a dataset of IT salaries. It "
                                                 "reads the provided dataset and performs"
                                                 "various statistical and aggregation operations on it. These "
                                                 "operations include computing the maximum,"
                                                 "minimum, median, and total salaries, as well as calculating the "
                                                 "average salary for a specific job and"
                                                 "experience level.")
    parser.add_argument('dataset', type=str, help='ds_salaries.csv')
    parser.add_argument('-o', '--output', type=str, help='the output Excel file name')
    args = parser.parse_args()
    dataset_name = args.dataset

    if not dataset_name.endswith('.csv'):
        sys.exit("Error: The file must have a .csv extension")

    if not os.path.exists(dataset_name):
        sys.exit("Error: The dataset file does not exist")

    args = parser.parse_args()
    dictionary = read_dataset(dataset_name)

    if args.output:
        check_file_exists(args.output, dictionary)
    else:
        generate_output(dictionary)


def read_dataset(filename):
    dataset_dict = {}
    try:
        with open(filename, 'r') as f:
            count = 0
            next(f)
            for line in f:
                pattern = r',\s*'
                fields = re.split(pattern, line)
                work_year, experience_level, employment_type, job_title, salary, salary_currency, \
                    salary_in_usd, employee_residence, remote_ratio, company_location, company_size = fields

                dataset_dict[f"{count}"] = {"work_year": work_year, "experience_level": experience_level,
                                            "employment_type": employment_type, "job_title": job_title,
                                            "salary": salary,
                                            "salary_currency": salary_currency, "salary_in_usd": salary_in_usd,
                                            "employee_residence": employee_residence,
                                            "remote_ratio": remote_ratio, "company_location": company_location,
                                            "company_size": company_size}

                count += 1

    except AssertionError as e:
        logger.error(e)

    return dataset_dict


def average_salary_by_job_experience(job_title, experience, dictionary):
    total, count = 0, 0
    for value in dictionary.values():
        if value["job_title"] == job_title \
                and value["experience_level"] == experience:
            total += int(value["salary"])
            count += 1

    if count != 0:
        return int(total / count)
    else:
        return "-"


def median_salary_by_job(job_title, dictionary):
    salary_list = []
    for value in dictionary.values():
        if value["job_title"] == job_title:
            salary_list.append(int(value["salary"]))

    sorted_salary = sorted(salary_list)
    middle_index = len(sorted_salary) // 2
    if len(sorted_salary) % 2 == 0:
        median = (sorted_salary[middle_index - 1] + sorted_salary[middle_index]) / 2
    else:
        median = sorted_salary[middle_index]

    return median


def min_salary_by_job_experience(job_title, experience, dictionary):
    min_val = float("inf")

    for value in dictionary.values():
        if value["job_title"] == job_title \
                and value["experience_level"] == experience:
            salary = float(value["salary"])
            if salary < min_val:
                min_val = salary

    if min_val == float("inf"):
        return "-"
    else:
        return int(min_val)


def max_salary_by_job_experience(job_title, experience, dictionary):
    max_val = float("-inf")

    for value in dictionary.values():
        if value["job_title"] == job_title \
                and value["experience_level"] == experience:
            salary = float(value["salary"])
            if salary > max_val:
                max_val = salary

    if max_val == float("-inf"):
        return "-"
    else:
        return int(max_val)


def max_salary(dictionary):
    max_val = float("-inf")

    for value in dictionary.values():
        salary = float(value["salary"])
        if salary > max_val:
            max_val = salary

    return int(max_val)


def min_salary(dictionary):
    min_val = float("inf")

    for value in dictionary.values():
        salary = float(value["salary"])
        if salary < min_val:
            min_val = salary

    return int(min_val)


def jobs(dictionary):
    unique_jobs = set()

    for value in dictionary.values():
        job_title = value["job_title"]
        unique_jobs.add(job_title)

    return unique_jobs


def experience_levels(dictionary):
    unique_experiences = set()

    for value in dictionary.values():
        experience = value["experience_level"]
        unique_experiences.add(experience)

    return unique_experiences


def num_of_countries(dictionary):
    unique_countries = set()

    for value in dictionary.values():
        country = value["company_location"]
        unique_countries.add(country)

    return len(unique_countries)


def check_file_exists(file_name, dictionary):
    counter = 1
    name, extension = os.path.splitext(file_name)

    while os.path.exists(file_name):
        file_name = f"{name}({counter}){extension}"
        counter += 1
    else:
        save_to_xls(file_name, dictionary)


def save_to_xls(file_name, dictionary):
    workbook = openpyxl.Workbook()
    worksheets = []
    worksheet1 = workbook.active
    worksheets.append(worksheet1)
    worksheet1.append(['Job Title', 'Experience Level', 'Average Salary', 'Max Salary', 'Min Salary'])

    for job in jobs(dictionary):
        for experience in experience_levels(dictionary):
            avg_salary = average_salary_by_job_experience(job, experience, dictionary)
            max_job_exp_salary = max_salary_by_job_experience(job, experience, dictionary)
            min_job_exp_salary = min_salary_by_job_experience(job, experience, dictionary)

            worksheet1.append([job, experience, avg_salary, max_job_exp_salary, min_job_exp_salary])

    worksheet2 = workbook.create_sheet("Sheet 2")
    worksheets.append(worksheet2)
    worksheet2.append(['Job Title', 'Median', 'Max Salary', 'Min Salary', 'Number of jobs', 'Number of countries'])

    for job in jobs(dictionary):
        median = median_salary_by_job(job, dictionary)
        worksheet2.append([job, median])

    worksheet2.cell(row=2, column=3).value = max_salary(dictionary)
    worksheet2.cell(row=2, column=4).value = min_salary(dictionary)
    worksheet2.cell(row=2, column=5).value = len(jobs(dictionary))
    worksheet2.cell(row=2, column=6).value = num_of_countries(dictionary)

    format_text(worksheets)

    workbook.save(file_name)

    output_file = os.path.abspath(file_name)
    print(f"File saved to {output_file}")


def format_text(worksheets):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')

    for worksheet in worksheets:
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        data_font = Font(color='000000')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font


def generate_output(dictionary):
    print("Average salary of EN Machine Learning Engineer ", average_salary_by_job_experience("Machine Learning "
                                                                                              "Engineer", "EN", dictionary))
    print("Median of salaries of Machine Learning Engineer: ", median_salary_by_job("Machine Learning Engineer", dictionary))
    print("Min salary of EN Machine Learning Engineer: ",
          min_salary_by_job_experience("Machine Learning Engineer", "EN", dictionary))
    print("Max salary of EN Machine Learning Engineer: ",
          max_salary_by_job_experience("Machine Learning Engineer", "EN", dictionary))
    print("Max salary: ", max_salary(dictionary))
    print("Min salary: ", min_salary(dictionary))
    print("Number of jobs: ", len(jobs(dictionary)))
    print("Number of countries: ", num_of_countries(dictionary))


run()
